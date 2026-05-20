import requests
from bs4 import BeautifulSoup
from datetime import date
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil

from getMitre_old_versions import get_mitre_old_version
from getMitre_new_versions import get_mitre_new_version
from changes_TacTech import cambiosTecnicasTacticas
from table_TacTech import generarTablasAsociacion


# Se solicita al usuario si quiere comprobar las diferencias entre dos versiones o si quiere obtener la última versión
opcion = input('''\nIntroduce el número correspondiente a la opción deseada:
    0 - OBTENER una versión y comprobar diferencias con otra versión
    1 - OBTENER una versión y NO comprobar diferencias con otra versión
    2 - SOLO comprobar diferencias entre versiones

    Opción: ''')

if opcion not in ['0', '1', '2']:
    print("Opción no válida. Por favor, vuelve a ejecutar el script e introduce 0, 1 o 2.")
    sys.exit(1)

if opcion == '0':
    excel_old = input("\nIntroduce el nombre del fichero Excel con la versión a comparar: ")
    excel_old = f'TechniquesTacticsMitre/{excel_old}'
    if not Path(excel_old).exists():
        print(f"No se encontró el fichero {excel_old} para comparar las diferencias.\nPor favor, vuelve a ejecutar el script e introduce un nombre válido.")
        sys.exit(1)

if opcion in ['0', '1']:
    # Se comprueba que existe el directorio y la plantilla
    excel_name = 'TechniquesTacticsMitre/TechniquesTactics.xlsx'
    excel_path = Path(excel_name)
    if not excel_path.exists():
        print(
            f"No se encontró el fichero base '{excel_name}'. "
            "Asegúrate de que existe para usarlo como plantilla."
        )
        sys.exit(1)

    # --- 1. RESPALDAR PLANTILLA ORIGINAL ---
    backup_name = 'TechniquesTacticsMitre/TechniquesTactics_Backup.xlsx'
    shutil.copy(excel_name, backup_name)

    # --- 2. VACIAR LA PLANTILLA POR COMPLETO (OPCIÓN NUCLEAR) ---
    # Esto asegura que no quede NI RASTRO de Enterprise antes de empezar
    wb_clean = openpyxl.load_workbook(excel_name)
    for sheet_name in wb_clean.sheetnames:
        ws_clean = wb_clean[sheet_name]
        if ws_clean.max_row > 1:
            # Borra todas las filas desde la fila 2 hacia abajo
            ws_clean.delete_rows(2, ws_clean.max_row)
    wb_clean.save(excel_name)
    wb_clean.close()

    # --- CAMBIO A MATRIZ ICS ---
    url = "https://attack.mitre.org/matrices/ics/"
    response = requests.get(url)
    
    # Se extrae la versión actual de la matriz de técnicas de MITRE
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        tag_version = soup.find("a", href=re.compile(r"^/versions/v\d+/matrices/ics/?$"))
        if tag_version:
            href = tag_version["href"]
            numero_version = re.search(r"/versions/(v\d+)/", href).group(1)
            version_actual = ["Versión MITRE ATT&CK:", numero_version]
    else:
        print("Error al realizar la solicitud:", response.status_code)
        sys.exit(1)    

    version_pedida = input(f'''\nIntroduce la versión de la matriz de técnicas de MITRE que deseas obtener (por ejemplo, v17):
Si quieres obtener la última versión, introduce 'latest' (actual es {numero_version}): ''')

    # Se valida el formato de la versión pedida
    if version_pedida != 'latest' and not re.match(r"^v\d+$", version_pedida):
        print("Formato de versión no válido.")
        sys.exit(1)

    # Gestión de versiones
    numero_version_actual = int(re.search(r"v(\d+)", numero_version).group(1))
    if version_pedida == 'latest':
        version_pedida = numero_version
        numero_version_pedida = numero_version_actual
    else:
        numero_version_pedida = int(re.search(r"v(\d+)", version_pedida).group(1))

    if version_pedida != 'latest' and (numero_version_pedida < 10 or numero_version_pedida > numero_version_actual):
        print(f"La versión {version_pedida} no está disponible.")
        sys.exit(1)

    if version_pedida != 'latest' and version_pedida != numero_version:
        url_version = f"https://attack.mitre.org/versions/{version_pedida}/matrices/ics/"
        response = requests.get(url_version)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            tag_version = soup.find("a", href=re.compile(r"^/versions/v\d+/matrices/ics/?$"))
            if tag_version:
                href = tag_version["href"]
                version = ["Versión MITRE ATT&CK:", re.search(r"/versions/(v\d+)/", href).group(1)]
        else:
            print(f"No se pudo obtener la versión {version_pedida} de ICS.")
            sys.exit(1)
    else:
        version = ["Versión MITRE ATT&CK:", version_pedida]

    contenedor = soup.find(class_='matrix-wrapper') or soup.find(id='matrix-container') or soup.find('table')
    
    if contenedor:
        tactic_links = contenedor.find_all('a', href=lambda href: href and '/tactics/T' in href)
        technique_links = contenedor.find_all('a', href=lambda href: href and '/techniques/T' in href)
    else:
        tactic_links = soup.find_all('a', href=lambda href: href and '/tactics/ics/T' in href)
        technique_links = soup.find_all('a', href=lambda href: href and '/techniques/ics/T' in href)

    # Diccionarios para Tácticas
    tactic_ids = []
    tactics = {}
    for link in tactic_links:
        tactic_id = link['href'].strip('/').split('/')[-1]
        tactic_name = link.text.strip()
        if tactic_id != '':
            tactics[tactic_id, "Tactic Name"] = tactic_name
            tactic_ids.append(tactic_id)

    # Diccionarios para Técnicas y Subtécnicas
    technique_ids = []
    subtechnique_ids = []
    techniques = {}
    subtechniques = {}
    
    for link in technique_links:
        clean_href = link['href'].strip('/')
        parts = clean_href.split('/')
        if not parts:
            continue
            
        technique_id = parts[-1]
        technique_name = link.text.strip()
        
        if technique_id.startswith('T'):
            techniques[technique_id, "Technique Name"] = technique_name
            if technique_id not in technique_ids:
                technique_ids.append(technique_id)
        elif technique_id.isnumeric() and len(parts) >= 2:
            esp_subtechnique_id = technique_id
            parent_id = parts[-2]
            
            if parent_id.startswith('T'):
                subtechnique_id = f"{parent_id}.{esp_subtechnique_id}"
                subtechniques[subtechnique_id, "Subtechnique Name"] = technique_name
                subtechniques[subtechnique_id, "Technique ID"] = parent_id
                if subtechnique_id not in subtechnique_ids:
                    subtechnique_ids.append(subtechnique_id)

    # Asociar nombre de técnica a subtécnica
    for sub_id in subtechnique_ids:
        parent_id = subtechniques[sub_id, "Technique ID"]
        if (parent_id, "Technique Name") in techniques:
            t_name = techniques[parent_id, "Technique Name"]
            if t_name.endswith(')'):
                t_name = t_name[:t_name.rfind('(')].strip()
            subtechniques[sub_id, "Technique Name"] = t_name
        else:
            subtechniques[sub_id, "Technique Name"] = "Unknown"

    # Limpieza y ordenación
    tactic_ids = sorted(list(set(tactic_ids)))
    technique_ids = sorted(list(set(technique_ids)))
    subtechnique_ids = sorted(list(set(subtechnique_ids)))

    for t_id in technique_ids:
        t_name = techniques[t_id, "Technique Name"]
        if t_name.endswith(')'):
            techniques[t_id, "Number of Subtechniques"] = t_name[t_name.rfind('(')+1:t_name.rfind(')')].strip()
            techniques[t_id, "Technique Name"] = t_name[:t_name.rfind('(')].strip()
        else:
            techniques[t_id, "Number of Subtechniques"] = 0

    # Fecha de actualización
    if version_pedida == 'latest' or version_pedida == numero_version:
        today = date.today()
        fecha = ["Fecha de la última actualización:", today.strftime("%d/%m/%Y")]
    else:
        banner = soup.select_one("div.version-banner")
        if banner:
            texto_fecha = banner.get_text(" ", strip=True)
            m = re.search(r'between\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\s+and\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})', texto_fecha)
            if m:
                inicio, fin = m.groups()
                fecha = ["Fechas de vigencia de la versión:", f"{inicio} - {fin}"]
            else:
                fecha = ["Fechas:", "No extraídas"]
        else:
            fecha = ["Fecha de la versión:", version_pedida]

    # Ejecución de los scripts que rellenan el Excel común (ahora totalmente vacío)
    if version_pedida in ['v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17']:
        get_mitre_old_version(tactics, techniques, subtechniques, tactic_ids, technique_ids, subtechnique_ids, version, fecha)
    else:
        get_mitre_new_version(tactics, techniques, subtechniques, tactic_ids, technique_ids, subtechnique_ids, version, fecha)
        
    generarTablasAsociacion()

    # --- 3. PROCESAMIENTO COMPLETO DEL NUEVO EXCEL (COLORES Y ANCHOS) ---
    wb = openpyxl.load_workbook(excel_name)
    
    # Paleta de colores: Verde Corporativo Elegante
    fill_header = PatternFill(start_color="2E5A44", end_color="2E5A44", fill_type="solid") # Verde oscuro
    fill_zebra = PatternFill(start_color="F2F7F4", end_color="F2F7F4", fill_type="solid")  # Verde clarito alterno
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11, color="000000")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Aplicar estilos a las cabeceras (Fila 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        
        # Aplicar estilos a los datos (Filas de ICS)
        if ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                is_zebra = (row % 2 == 0)
                current_fill = fill_zebra if is_zebra else fill_white
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = current_fill
                    cell.font = font_body
                    cell.border = border_all
                    
                    # Alineación
                    header_val = str(ws.cell(row=1, column=col).value or "")
                    if col == 1 or "ID" in header_val:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
        # Autoajustar el ancho de las columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Guardamos el resultado final con el nombre limpio de la versión
    nuevo_nombre = f'TechniquesTacticsMitre/ICS_{version_pedida}.xlsx'
    wb.save(nuevo_nombre)
    wb.close()
    
    # --- 4. RESTAURAR LA PLANTILLA ORIGINAL ---
    shutil.move(backup_name, excel_name)
    
    print(f"\n¡Perfecto! Documento completado: {nuevo_nombre}")
    print("El archivo ha sido purgado completamente antes de la extracción y tiene aplicado el formato verde.")

    if opcion == '0':
        cambiosTecnicasTacticas(nuevo_nombre, excel_old)

if opcion == '2':
    excel_n1 = input("\nIntroduce el nombre del primer fichero Excel: ")
    excel_n1 = f'TechniquesTacticsMitre/{excel_n1}'
    excel_n2 = input("Introduce el nombre del segundo fichero Excel: ")
    excel_n2 = f'TechniquesTacticsMitre/{excel_n2}'
    cambiosTecnicasTacticas(excel_n1, excel_n2)